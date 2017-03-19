#! python3
# sendDuesReminders.py - Sends emails based on their status in spreadsheet.

import openpyxl, smtplib, sys
import datetime

whoami = 'gavin.low@greatwhiteproductionsltd.com'
date = datetime.datetime.now().strftime( "%d/%m/%Y %H:%M" )
# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Corporate')

lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol).value

unsentMembers = {}
# Check each member's payment status
for r in range(2, sheet.max_row + 1):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'sent':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unsentMembers[name] = email

# Log in to email account.
smtpObj = smtplib.SMTP_SSL('smtp.zoho.com', 465)
#smtpObj.ehlo()
#smtpObj.starttls()
#enter the password as a command line argument each time you run the program, to avoid saving your password in your source code.
smtpObj.login(whoami, '***')

# Send out reminder emails.
for name, email in unsentMembers.items():
    body = 'From:%s\nTo:%s\nDate:%s\nSubject: Corporate Video.\nDear All,\nGood day,\nGreetings from Great White Productions. Hope this email finds you well.\n\nGreat White Productions offer a wide range of professional media production services such as commercial shoots , advertising video , opening ceremony ,and many more.\n\nOur Professional team were invited to Telunas resorts in batam to capture the 10th year anniversary and the opening of the new resort.\n\nShould you require any production services , do feel free to contact us for more details.\n\n\nRegards,\nGavin Low\nBusiness Development Director\nGreat White Productions Ltd.\n140 Paya Lebar Road #10-12 Singapore 409015\nEmail: gavin.low@greatwhiteproductionsltd.com\nWebsite: www.greatwhiteproductionsltd.com' % ( whoami,email,date )
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail(whoami, email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))
smtpObj.quit()
